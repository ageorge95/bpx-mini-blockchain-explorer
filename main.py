from web3 import Web3,\
    HTTPProvider
from json import load,\
    dump
from os import path
from openpyxl import Workbook

# ######################################
# USER INPUT
# ######################################
w3 = Web3(HTTPProvider('http://127.0.0.1:8545'))
blockchain_address = input('What ethereum address should I scan ?\n\t_')

# ######################################
# SCRIPT
# ######################################
# load the cached data
if path.isfile('data.json'):
    try:
        with open('data.json', 'r') as input_file_handle:
            tx_dictionary = load(input_file_handle)
    except:
        print('Failed to load data.json. Bootstraping a new cache file ...')
        tx_dictionary = {}
else:
    tx_dictionary = {}

# check if this address was used in the past
if blockchain_address not in tx_dictionary.keys():
    tx_dictionary[blockchain_address] = []

# input_data
latest_blocknumber = w3.eth.block_number
denominator = 10**9

for block_number in range(latest_blocknumber):
    print(f'Analyzing transactions from block {block_number}')
    if (not tx_dictionary[blockchain_address]) or\
            (block_number > tx_dictionary[blockchain_address][-1]['block_number']):
        tx_dictionary[blockchain_address].append({'block_number': block_number,
                                                  'transactions': [],
                                                  'withdrawals': []})
        block = w3.eth.get_block(block_number, True)
        transactions_found = 0
        for transaction in block.transactions:
            if transaction['to'] == blockchain_address:
                tx_dictionary[blockchain_address][-1]['transactions'].append({'from': transaction['from'],
                                                                              'to': transaction['to'],
                                                                              'value': transaction['value']})
                transactions_found += 1
            if transaction['from'] == blockchain_address:
                tx_dictionary[blockchain_address][-1]['transactions'].append({'from': transaction['from'],
                                                                              'to': transaction['to'],
                                                                              'value': -transaction['value']})
                transactions_found += 1
        withdrawals_found = 0
        for withdrawal in block.withdrawals:
            if withdrawal.address == blockchain_address:
                tx_dictionary[blockchain_address][-1]['withdrawals'].append({'amount': withdrawal.amount})
                withdrawals_found == 1

        print(f'Block cached; Found {transactions_found} transactions and {withdrawals_found} withdrawals')
    else:
        print('Block already cached')

with open('data.json', 'w') as output_file_handle:
    dump(tx_dictionary, output_file_handle, indent=2)

# prepare the tabular data
tx_dictionary[blockchain_address].reverse()

table_data = []
for data in tx_dictionary[blockchain_address]:
    if data['transactions']:
        for transaction in data['transactions']:
            table_data.append([data['block_number'],
                               'transaction',
                               transaction['from'],
                               transaction['to'],
                               transaction['value']/denominator/denominator])
    if data['withdrawals']:
        for withdrawal in data['withdrawals']:
            table_data.append([data['block_number'],
                               'withdrawal',
                               None,
                               None,
                               withdrawal['amount']/denominator])

# create and save an excel export
wb = Workbook()
ws = wb['Sheet']

ws[f'A1'].value = 'Block_number'
ws[f'B1'].value = 'Type'
ws[f'C1'].value = 'From'
ws[f'D1'].value = 'To'
ws[f'E1'].value = 'Amount'

for i, _ in enumerate(table_data, 2):
    ws[f'A{i}'].value = _[0]
    ws[f'B{i}'].value = _[1]
    ws[f'C{i}'].value = _[2]
    ws[f'D{i}'].value = _[3]
    ws[f'E{i}'].value = _[4]

wb.save('export.xlsx')